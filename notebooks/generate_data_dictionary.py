"""
FreightFlow NextGen Analytics
Data Dictionary Generator — Silver Layer

Reads all Silver Parquet tables and generates a comprehensive
Data Dictionary Excel file in docs/data-dictionary/

Run:
    python notebooks/generate_data_dictionary.py
"""

from pyspark.sql import SparkSession
from pyspark.sql import functions as F
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Spark Session ─────────────────────────────────────────────────────────────
spark = SparkSession.builder \
    .appName("FreightFlow_Data_Dictionary") \
    .getOrCreate()
spark.sparkContext.setLogLevel("WARN")

SILVER_PATH = "data/silver/"
OUTPUT_PATH = "docs/data-dictionary/Data_Dictionary_FreightFlow.xlsx"
os.makedirs("docs/data-dictionary", exist_ok=True)

# ── Column Descriptions (from STT Mapping) ────────────────────────────────────
descriptions = {
    "vehicle": {
        "vehicle_id":           "Primary key. Unique vehicle identifier. Shared parent key for Trucks and Trailers.",
        "registration_number":  "Vehicle registration number. Unique constraint retained from source.",
        "last_service_date":    "Date of last maintenance or service event. Null if never serviced.",
        "next_service_date":    "Scheduled date of next service event.",
        "vehicle_status":       "Current operational status. Values: Active, Dormant, Parked, UnderMaintenance.",
        "vehicle_type":         "Type of vehicle. Values: Truck, Trailer.",
    },
    "trucks": {
        "truck_id":             "Primary key. FK to vehicle.vehicle_id.",
        "license_plate":        "Vehicle license plate number. Uppercased.",
        "engine_type":          "Engine fuel type. Values: Diesel, Gasoline, Electric.",
        "current_mileage_k":    "Current odometer reading in thousands of miles. Must be >= 0.",
        "vehicle_id":           "Foreign key referencing vehicle.vehicle_id.",
    },
    "trailers": {
        "trailer_id":           "Primary key. FK to vehicle.vehicle_id.",
        "capacity_lbs":         "Maximum load capacity in pounds. Must be > 0.",
        "trailer_type":         "Trailer body type. Values: Flatbed, Dryvan, Lowboy, Stepdeck.",
        "has_refrigeration":    "Whether trailer has refrigeration capability. Boolean.",
        "vehicle_id":           "Foreign key referencing vehicle.vehicle_id.",
    },
    "service_event": {
        "service_event_id":     "Primary key. Unique service event identifier.",
        "service_date":         "Date and time the service event occurred.",
        "service_type":         "Type of service performed. Values: Repair, Maintenance, Tires, Replace_Part.",
        "event_description":    "Free-text description of the service event.",
        "outcome":              "Result of the service event. Values: Ready_for_Dispatch, Under_Maintenance.",
        "mechanic_cost_usd":    "Cost of mechanic labor and parts in USD. Must be >= 0.",
        "vehicle_id":           "Foreign key referencing vehicle.vehicle_id.",
    },
    "driver": {
        "driver_id":            "Primary key. Unique driver identifier.",
        "first_name":           "Driver first name. Trimmed.",
        "last_name":            "Driver last name. Trimmed.",
        "license_number":       "Commercial driver license number. Uppercased.",
        "license_expiry_date":  "Expiry date of driver license. Date only (time component dropped).",
        "hos_status":           "Hours of Service compliance status. Values: Compliant, Non-Compliant.",
        "phone":                "Driver contact phone number.",
        "email":                "Driver email address. Lowercased.",
        "address":              "Driver street address.",
        "city":                 "Driver city. Title-cased.",
        "state":                "Driver state.",
        "zip_code":             "Driver ZIP code. 5-digit integer.",
        "availability":         "Current availability status. Values: Available, Unavailable.",
    },
    "client": {
        "client_id":            "Primary key. Unique client identifier.",
        "company_name":         "Client company name.",
        "point_of_contact":     "Name of the primary contact person at the client.",
        "phone":                "Client contact phone number.",
        "email":                "Client email address. Lowercased.",
        "billing_address":      "Client billing address.",
    },
    "third_party_carrier": {
        "carrier_id":           "Primary key. Unique carrier identifier.",
        "carrier_name":         "Name of the third-party carrier company.",
        "point_of_contact":     "Name of the primary contact at the carrier.",
        "phone":                "Carrier contact phone number.",
        "email":                "Carrier email address. Lowercased.",
        "compliance_status":    "Regulatory compliance status. Values: Compliant, Non-Compliant.",
    },
    "tender": {
        "tender_id":            "Primary key. Unique tender/order identifier.",
        "product_description":  "Description of the product being shipped.",
        "pickup_location":      "Pickup address for the tender.",
        "dropoff_location":     "Delivery address for the tender.",
        "expected_delivery_date": "Expected date and time of delivery.",
        "temp_requirements":    "Temperature control requirement. Values: Ambient, Frozen, Deep Frozen, Controlled Cool.",
        "weight_lbs":           "Shipment weight in pounds. Must be > 0.",
        "distance_miles":       "Estimated shipment distance in miles. Must be > 0.",
        "client_id":            "Foreign key referencing client.client_id.",
    },
    "shipment": {
        "shipment_id":          "Primary key. Unique shipment identifier.",
        "pickup_point":         "Actual pickup address for the shipment.",
        "delivery_point":       "Actual delivery address for the shipment.",
        "delivery_status":      "Current delivery status. Values: Delivered, Pending, In_Transit.",
        "delivery_date":        "Actual or scheduled delivery date and time.",
        "temp_requirements":    "Temperature control requirement. Cross-validated against tender.temp_requirements.",
        "assigned_driver_id":   "Driver assigned to the shipment. Nullable — may not be assigned yet.",
        "distance_miles":       "Actual shipment distance in miles. Must be > 0.",
        "tender_id":            "Foreign key referencing tender.tender_id.",
    },
    "internal_shipment": {
        "internal_shipment_id": "Primary key and foreign key referencing shipment.shipment_id.",
        "fuel_cost_usd":        "Fuel cost for internal shipment in USD. Cast from INT to DECIMAL. Must be >= 0.",
    },
    "brokered_shipment": {
        "brokered_shipment_id": "Primary key and foreign key referencing shipment.shipment_id.",
        "brokered_status":      "Status of the brokered shipment. Values: Delivered, Pending.",
        "profit_margin_usd":    "Profit margin on the brokered shipment in USD. Must be >= 0.",
        "rate_usd":             "Rate charged for the brokered shipment in USD. Must be > 0.",
        "carrier_id":           "Foreign key referencing third_party_carrier.carrier_id.",
    },
    "dispatch": {
        "shipment_id":          "Part of composite PK. FK referencing internal_shipment.internal_shipment_id.",
        "driver_id":            "Part of composite PK. FK referencing driver.driver_id.",
        "truck_id":             "Part of composite PK. FK referencing trucks.truck_id.",
        "trailer_id":           "FK referencing trailers.trailer_id. Nullable per relational model.",
    },
}

# ── Excel Styles ──────────────────────────────────────────────────────────────
BLUE_DARK  = "1F4E79"
BLUE_MID   = "2E75B6"
BLUE_LIGHT = "BDD7EE"
WHITE      = "FFFFFF"
GRAY       = "F2F2F2"
ORANGE     = "FCE4D6"
GREEN      = "E2EFDA"
YELLOW     = "FFF2CC"

def xfill(c):  return PatternFill("solid", fgColor=c)
def xleft():   return Alignment(horizontal="left", vertical="center", wrap_text=True)
def xcenter(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
xthin = Side(style="thin", color="AAAAAA")
def xbthin(): return Border(left=xthin, right=xthin, top=xthin, bottom=xthin)

def xsh(cell, bg=BLUE_DARK):
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill      = xfill(bg)
    cell.alignment = xcenter()
    cell.border    = xbthin()

def xsb(cell, bg=WHITE, bold=False, c=False):
    cell.font      = Font(name="Arial", size=10, bold=bold)
    cell.fill      = xfill(bg)
    cell.alignment = xcenter() if c else xleft()
    cell.border    = xbthin()

# ── Build Workbook ────────────────────────────────────────────────────────────
wb = Workbook()
default = wb.active
sheet_map = {}

table_colors = {
    "vehicle":             "DEEAF1",
    "trucks":              "E2EFDA",
    "trailers":            "FFF2CC",
    "service_event":       "FCE4D6",
    "driver":              "EAD1DC",
    "client":              "D9EAD3",
    "third_party_carrier": "CFE2F3",
    "tender":              "FDE9D9",
    "shipment":            "D9D9F3",
    "internal_shipment":   "E6F2F8",
    "brokered_shipment":   "F4CCCC",
    "dispatch":            "DDEBF7",
}

headers    = ["Column Name", "Data Type", "Nullable", "Null Count", "Distinct Count", "Sample Values", "Description"]
col_widths = [28, 16, 10, 12, 15, 38, 55]

tables = list(descriptions.keys())

for table_name in tables:
    print(f"Processing: {table_name}")
    path = os.path.join(SILVER_PATH, table_name)
    df   = spark.read.parquet(path)
    total_rows = df.count()
    color = table_colors.get(table_name, WHITE)

    ws = wb.create_sheet(table_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Title banner
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    t = ws.cell(1, 1, f"Data Dictionary — {table_name}  |  Silver Layer  |  {total_rows} rows")
    t.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t.fill      = xfill(BLUE_DARK)
    t.alignment = xcenter()
    t.border    = xbthin()
    ws.row_dimensions[1].height = 24

    # Headers
    for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(2, i, h)
        xsh(c, bg=BLUE_MID)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 20

    # Per-column stats
    col_stats = {}
    for col in df.columns:
        null_count     = df.filter(F.col(col).isNull()).count()
        distinct_count = df.select(col).distinct().count()
        # Sample: up to 3 non-null distinct values
        samples = df.filter(F.col(col).isNotNull()) \
                    .select(col).distinct().limit(3) \
                    .rdd.flatMap(lambda x: x).collect()
        sample_str = ", ".join(str(s) for s in samples)
        col_stats[col] = (null_count, distinct_count, sample_str)

    for r_idx, col in enumerate(df.columns, start=3):
        bg = color if r_idx % 2 == 1 else WHITE
        null_count, distinct_count, sample_str = col_stats[col]
        dtype       = str(df.schema[col].dataType)
        nullable    = "Yes" if df.schema[col].nullable else "No"
        desc        = descriptions.get(table_name, {}).get(col, "")

        ws.row_dimensions[r_idx].height = 18

        xsb(ws.cell(r_idx, 1, col),            bg=bg, bold=True)
        xsb(ws.cell(r_idx, 2, dtype),          bg=bg, c=True)
        xsb(ws.cell(r_idx, 3, nullable),       bg=bg, c=True)

        # Null count — highlight if any nulls
        nc = ws.cell(r_idx, 4, null_count)
        nc_bg = ORANGE if null_count > 0 else bg
        xsb(nc, bg=nc_bg, c=True)

        xsb(ws.cell(r_idx, 5, distinct_count), bg=bg, c=True)
        xsb(ws.cell(r_idx, 6, sample_str),     bg=bg)
        xsb(ws.cell(r_idx, 7, desc),           bg=bg)

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(df.columns)+2}"
    sheet_map[table_name] = ws

    # Back link
    last = len(df.columns) + 4
    back = ws.cell(last, 1, "← Back to Index")
    back.font      = Font(name="Arial", size=9, color=BLUE_MID, underline="single", italic=True)
    back.alignment = xleft()
    back.hyperlink = "#'Data Dictionary Index'!B2"

# ── INDEX SHEET ───────────────────────────────────────────────────────────────
ws_idx = wb.create_sheet("Data Dictionary Index", 0)
ws_idx.sheet_view.showGridLines = False

ws_idx.merge_cells("B2:G2")
ws_idx.cell(2, 2, "FreightFlow NextGen Analytics — Data Dictionary (Silver Layer)").font = \
    Font(name="Arial", bold=True, size=16, color=BLUE_DARK)
ws_idx.cell(2, 2).alignment = Alignment(horizontal="left", vertical="center")
ws_idx.row_dimensions[2].height = 32

ws_idx.merge_cells("B3:G3")
ws_idx.cell(3, 2, "Auto-generated from Silver Parquet files  |  Version 1.0  |  March 2026  |  Click table name to navigate").font = \
    Font(name="Arial", size=10, color="888888", italic=True)
ws_idx.cell(3, 2).alignment = Alignment(horizontal="left", vertical="center")
ws_idx.row_dimensions[3].height = 20

idx_hdrs   = ["#", "Table Name", "Row Count", "Column Count", "Nullable Columns", "Non-Nullable Columns"]
idx_widths = [5, 28, 12, 14, 18, 20]
for i, (h, w) in enumerate(zip(idx_hdrs, idx_widths)):
    col = i + 2
    c = ws_idx.cell(5, col, h)
    xsh(c, bg=BLUE_DARK)
    ws_idx.column_dimensions[get_column_letter(col)].width = w
ws_idx.row_dimensions[5].height = 22

for i, table_name in enumerate(tables, start=1):
    row = i + 5
    ws_idx.row_dimensions[row].height = 22
    bg   = table_colors.get(table_name, WHITE)
    path = os.path.join(SILVER_PATH, table_name)
    df   = spark.read.parquet(path)

    row_count      = df.count()
    col_count      = len(df.columns)
    nullable_cols  = sum(1 for f in df.schema.fields if f.nullable)
    non_null_cols  = col_count - nullable_cols

    xsb(ws_idx.cell(row, 2, i), bg=bg, c=True)

    c_name = ws_idx.cell(row, 3, table_name)
    c_name.font      = Font(name="Arial", size=10, bold=True, color=BLUE_MID, underline="single")
    c_name.fill      = xfill(bg)
    c_name.alignment = xleft()
    c_name.border    = xbthin()
    c_name.hyperlink = f"#{table_name}!A1"

    xsb(ws_idx.cell(row, 4, row_count),     bg=bg, c=True)
    xsb(ws_idx.cell(row, 5, col_count),     bg=bg, c=True)
    xsb(ws_idx.cell(row, 6, nullable_cols), bg=bg, c=True)
    xsb(ws_idx.cell(row, 7, non_null_cols), bg=bg, c=True)

wb.remove(default)
wb.save(OUTPUT_PATH)
print(f"\n✅ Data Dictionary saved to: {OUTPUT_PATH}")
spark.stop()
